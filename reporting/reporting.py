from openpyxl import Workbook
import os.path

def generate_report(path, filename):
    '''
    
    '''

    # Report spec

    # Basic Workbook layout:
    #########################
    # 5 Sheets:
    #############
    # 1. PortfolioSummary
    # 2. AssetSummary
    # 3. CorrelationMatrix
    # 4. Leverage and Asset Allocation
    # 5. Exposure
    #
    #
    sheets = ['PortfolioSumamry',
              'AssetSummary',
              'CorrelationMatrix', 
              'Leverage and Asset Allocation',
              'Exposure'
              ]
    
    # create workbook
    wb = Workbook()
    
    # create the new sheets:
    # get active sheet and rename
    ps = wb.active
    ps.title = sheets[0]
    #
    
    for sheet in sheets[1:]
        ass = wb.create_sheet()
        ass.title = sheet
        
        cm = wb.create_sheet()
        cm.title = sheet
        
        laa = wb.create_sheet()
        laa.title = sheet
        
        exp = wb.create_sheet()
        exp.title = sheet
    
    wb.save(os.path.join(path, filename))